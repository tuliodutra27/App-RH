"""
Comparador de planilhas do RH - AliseoSA
=========================================

Compara a planilha do RH mais recente (relação-DD-MM.xlsx) com o
último snapshot salvo e gera/atualiza um arquivo de histórico
consolidado com adições, remoções e alterações de cargo,
departamento e gestor.

Tratamento de PJ: como colaboradores PJ não possuem matrícula
(a coluna vem preenchida apenas com "PJ"), eles são identificados
pelo Nome. CLTs são identificados pela Matrícula normalmente.

Uso:
    python comparador_rh.py "C:/caminho/relação-14-05.xlsx"

Ou rode sem argumentos e ele pega o arquivo "relação-*.xlsx" mais
recente da pasta_rh configurada abaixo.
"""

import sys
import re
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================
# CONFIG - ajuste estes caminhos para o seu ambiente
# ============================================================
CONFIG = {
    # Pasta onde o RH deposita os arquivos "relação-DD-MM.xlsx"
    "pasta_rh": r"planilha_rh",

    # Pasta separada para o histórico (snapshots + consolidado)
    "pasta_historico": r"old",

    # Nome do arquivo consolidado de histórico
    "arquivo_consolidado": "historico_alteracoes.xlsx",

    # Nomes das colunas na planilha do RH (após normalização para MAIÚSCULAS)
    "col_matricula":    "MATRÍCULA",
    "col_nome":         "NOME",
    "col_cargo":        "CARGO",
    "col_departamento": "DEPARTAMENTO",
    "col_gestor":       "GESTOR",
    "col_admissao":     "ADMISSÃO",
}

# Campos monitorados para detectar "ALTERAÇÕES"
CAMPOS_MONITORADOS = ["CARGO", "DEPARTAMENTO", "GESTOR"]


# ============================================================
# Estilos visuais
# ============================================================
FONT_TITLE     = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_SECTION   = Font(name="Arial", size=12, bold=True, color="FFFFFF")
FONT_HEADER    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_BODY      = Font(name="Arial", size=10)
FONT_BODY_BOLD = Font(name="Arial", size=10, bold=True)

FILL_TITLE  = PatternFill("solid", start_color="263238")
FILL_ADD    = PatternFill("solid", start_color="2E7D32")  # verde
FILL_REM    = PatternFill("solid", start_color="C62828")  # vermelho
FILL_ALT    = PatternFill("solid", start_color="EF6C00")  # laranja
FILL_HEADER = PatternFill("solid", start_color="455A64")  # cinza-azulado
FILL_ZEBRA  = PatternFill("solid", start_color="ECEFF1")  # cinza claro
FILL_DIFF   = PatternFill("solid", start_color="FFF59D")  # amarelo - célula alterada

THIN = Side(border_style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ============================================================
# Carga e normalização
# ============================================================
def carregar_planilha(caminho: Path) -> pd.DataFrame:
    """Carrega e normaliza uma planilha do RH."""
    df = pd.read_excel(caminho, dtype=str)
    df.columns = [str(c).strip().upper() for c in df.columns]

    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"NAN": "", "NONE": "", "NAT": ""})

    # Descarta linhas em branco (sem matrícula E sem nome) — totais, rodapés, etc.
    col_m = CONFIG["col_matricula"]
    col_n = CONFIG["col_nome"]
    antes = len(df)
    df = df[~((df[col_m] == "") & (df[col_n] == ""))].copy()
    descartadas = antes - len(df)
    if descartadas:
        print(f"   ℹ️  {descartadas} linha(s) em branco descartadas")

    # Chave: matrícula limpa para CLT, "PJ::<nome>" para PJ
    def gerar_chave(row):
        matr_raw = str(row.get(col_m, "")).strip()
        nome = str(row.get(col_n, "")).strip().upper()

        # Remove .0 que pandas adiciona quando lê número como str: "610.0" -> "610"
        matr = matr_raw.upper()
        if matr.endswith(".0") and matr[:-2].replace("-", "").isdigit():
            matr = matr[:-2]

        if matr in ("PJ", "", "NAN", "NONE"):
            if not nome:
                # Sem matrícula útil e sem nome -> chave aleatória pra não colidir
                return f"VAZIA::{id(row)}"
            return f"PJ::{nome}"
        return f"CLT::{matr}"

    df["_CHAVE"] = df.apply(gerar_chave, axis=1)

    # Limpa também a matrícula visível pra não aparecer "610.0" no relatório
    def limpar_matr(v):
        v = str(v).strip()
        if v.endswith(".0") and v[:-2].replace("-", "").isdigit():
            return v[:-2]
        return v
    df[col_m] = df[col_m].apply(limpar_matr)

    dup = df[df["_CHAVE"].duplicated(keep=False)].sort_values("_CHAVE")
    if not dup.empty:
        print(f"\n⚠️  ATENÇÃO: {len(dup)} linhas com chave duplicada em {caminho.name}:")
        for _, row in dup.iterrows():
            print(f"    {row['_CHAVE']} - {row.get(col_n, '')}")
        print("    (Mantendo a última ocorrência de cada chave)\n")
        df = df.drop_duplicates(subset="_CHAVE", keep="last")

    return df.set_index("_CHAVE")


# ============================================================
# Comparação
# ============================================================
def comparar(df_antigo: pd.DataFrame, df_novo: pd.DataFrame):
    """Retorna (adicoes, remocoes, alteracoes) como DataFrames."""
    k_ant, k_nov = set(df_antigo.index), set(df_novo.index)

    adicoes  = df_novo.loc[list(k_nov - k_ant)].copy()
    remocoes = df_antigo.loc[list(k_ant - k_nov)].copy()

    linhas_alt = []
    for chave in (k_ant & k_nov):
        antigo, novo = df_antigo.loc[chave], df_novo.loc[chave]
        diffs = {}
        for campo in CAMPOS_MONITORADOS:
            v_ant = str(antigo.get(campo, "")).strip()
            v_nov = str(novo.get(campo, "")).strip()
            if v_ant != v_nov:
                diffs[campo] = (v_ant, v_nov)
        if diffs:
            linha = {
                "TIPO": "PJ" if chave.startswith("PJ::") else "CLT",
                "MATRÍCULA": novo.get(CONFIG["col_matricula"], ""),
                "NOME": novo.get(CONFIG["col_nome"], ""),
            }
            for campo in CAMPOS_MONITORADOS:
                if campo in diffs:
                    linha[f"{campo} ANTES"]  = diffs[campo][0]
                    linha[f"{campo} DEPOIS"] = diffs[campo][1]
                else:
                    v = novo.get(campo, "")
                    linha[f"{campo} ANTES"]  = v
                    linha[f"{campo} DEPOIS"] = v
            linha["_CAMPOS_ALTERADOS"] = ", ".join(diffs.keys())
            linhas_alt.append(linha)

    alteracoes = pd.DataFrame(linhas_alt)

    for df in (adicoes, remocoes):
        if not df.empty:
            df.insert(0, "TIPO",
                      ["PJ" if k.startswith("PJ::") else "CLT" for k in df.index])

    return adicoes, remocoes, alteracoes


# ============================================================
# Escrita do relatório
# ============================================================
def proximo_nome_aba(wb: Workbook, data_str: str) -> str:
    base = f"Alterações {data_str}"
    if base not in wb.sheetnames:
        return base
    i = 2
    while f"{base} ({i})" in wb.sheetnames:
        i += 1
    return f"{base} ({i})"


def _escrever_secao(ws, linha, titulo, fill_titulo, df, colunas):
    """Adições ou remoções."""
    ws.merge_cells(start_row=linha, start_column=1,
                   end_row=linha,   end_column=len(colunas))
    c = ws.cell(row=linha, column=1, value=titulo)
    c.font = FONT_SECTION
    c.fill = fill_titulo
    c.alignment = LEFT
    ws.row_dimensions[linha].height = 22
    linha += 1

    if df.empty:
        c = ws.cell(row=linha, column=1, value="Nenhuma ocorrência.")
        c.font = Font(name="Arial", size=10, italic=True, color="78909C")
        ws.merge_cells(start_row=linha, start_column=1,
                       end_row=linha,   end_column=len(colunas))
        return linha + 2

    for idx, col in enumerate(colunas, start=1):
        c = ws.cell(row=linha, column=idx, value=col)
        c.font, c.fill, c.alignment, c.border = FONT_HEADER, FILL_HEADER, CENTER, BORDER
    linha += 1

    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(colunas, start=1):
            valor = row.get(col, "")
            c = ws.cell(row=linha, column=j,
                        value=str(valor) if pd.notna(valor) else "")
            c.font = FONT_BODY
            c.alignment = LEFT if j > 2 else CENTER
            c.border = BORDER
            if i % 2 == 1:
                c.fill = FILL_ZEBRA
        linha += 1

    return linha + 1


def _escrever_alteracoes(ws, linha, df):
    """Tabela ANTES vs DEPOIS para alterações."""
    colunas = ["TIPO", "MATRÍCULA", "NOME"]
    for campo in CAMPOS_MONITORADOS:
        colunas += [f"{campo} ANTES", f"{campo} DEPOIS"]
    total = len(colunas)

    ws.merge_cells(start_row=linha, start_column=1,
                   end_row=linha,   end_column=total)
    c = ws.cell(row=linha, column=1, value=f"✏️  ALTERAÇÕES ({len(df)})")
    c.font, c.fill, c.alignment = FONT_SECTION, FILL_ALT, LEFT
    ws.row_dimensions[linha].height = 22
    linha += 1

    if df.empty:
        c = ws.cell(row=linha, column=1, value="Nenhuma ocorrência.")
        c.font = Font(name="Arial", size=10, italic=True, color="78909C")
        ws.merge_cells(start_row=linha, start_column=1,
                       end_row=linha,   end_column=total)
        return linha + 2

    for idx, col in enumerate(colunas, start=1):
        c = ws.cell(row=linha, column=idx, value=col)
        c.font, c.fill, c.alignment, c.border = FONT_HEADER, FILL_HEADER, CENTER, BORDER
    linha += 1

    for i, (_, row) in enumerate(df.iterrows()):
        campos_alt = set((row.get("_CAMPOS_ALTERADOS") or "").split(", "))
        for j, col in enumerate(colunas, start=1):
            valor = row.get(col, "")
            c = ws.cell(row=linha, column=j,
                        value=str(valor) if pd.notna(valor) else "")
            c.font = FONT_BODY
            c.alignment = LEFT if j > 2 else CENTER
            c.border = BORDER
            if i % 2 == 1:
                c.fill = FILL_ZEBRA
            # Destaca em amarelo as colunas dos campos que de fato mudaram
            for campo in campos_alt:
                if col in (f"{campo} ANTES", f"{campo} DEPOIS"):
                    c.fill = FILL_DIFF
                    c.font = FONT_BODY_BOLD
        linha += 1

    return linha + 1


def escrever_aba(wb, data_str, arquivo_origem, adicoes, remocoes, alteracoes):
    nome_aba = proximo_nome_aba(wb, data_str)
    ws = wb.create_sheet(nome_aba)

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"Relatório de Alterações - {data_str}"
    c.font, c.fill, c.alignment = FONT_TITLE, FILL_TITLE, CENTER
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Arquivo origem: {arquivo_origem}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="546E7A")
    ws.merge_cells("A2:I2")

    ws["A3"] = (f"Resumo: {len(adicoes)} adições  |  "
                f"{len(remocoes)} remoções  |  {len(alteracoes)} alterações")
    ws["A3"].font = FONT_BODY_BOLD
    ws.merge_cells("A3:I3")

    linha = 5
    linha = _escrever_secao(
        ws, linha,
        f"➕ ADIÇÕES ({len(adicoes)})", FILL_ADD, adicoes,
        ["TIPO", CONFIG["col_matricula"], CONFIG["col_nome"],
         CONFIG["col_cargo"], CONFIG["col_departamento"],
         CONFIG["col_gestor"], CONFIG["col_admissao"]],
    )
    linha = _escrever_secao(
        ws, linha,
        f"➖ REMOÇÕES ({len(remocoes)})", FILL_REM, remocoes,
        ["TIPO", CONFIG["col_matricula"], CONFIG["col_nome"],
         CONFIG["col_cargo"], CONFIG["col_departamento"], CONFIG["col_gestor"]],
    )
    _escrever_alteracoes(ws, linha, alteracoes)

    larguras = {"A": 8, "B": 12, "C": 28, "D": 30, "E": 28,
                "F": 14, "G": 14, "H": 14, "I": 14}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


def atualizar_indice(wb: Workbook):
    """Recria a aba 'Índice' listando todas as abas de alterações."""
    if "Índice" in wb.sheetnames:
        del wb["Índice"]
    idx = wb.create_sheet("Índice", 0)

    idx.merge_cells("A1:C1")
    c = idx["A1"]
    c.value = "Histórico de Alterações - RH AliseoSA"
    c.font, c.fill, c.alignment = FONT_TITLE, FILL_TITLE, CENTER
    idx.row_dimensions[1].height = 28

    for j, h in enumerate(["Aba", "Data do relatório", "Última atualização"], start=1):
        c = idx.cell(row=3, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = FONT_HEADER, FILL_HEADER, CENTER, BORDER

    linha = 4
    for nome in wb.sheetnames:
        if nome == "Índice":
            continue
        idx.cell(row=linha, column=1, value=nome).font = FONT_BODY
        m = re.search(r"(\d{2}-\d{2}(?:-\d{4})?)", nome)
        idx.cell(row=linha, column=2, value=m.group(1) if m else "").font = FONT_BODY
        idx.cell(row=linha, column=3,
                 value=datetime.now().strftime("%d/%m/%Y %H:%M")).font = FONT_BODY
        for j in range(1, 4):
            idx.cell(row=linha, column=j).border = BORDER
            idx.cell(row=linha, column=j).alignment = LEFT
        linha += 1

    idx.column_dimensions["A"].width = 35
    idx.column_dimensions["B"].width = 22
    idx.column_dimensions["C"].width = 22


# ============================================================
# Orquestração
# ============================================================
def resolver_arquivo_novo() -> Path:
    if len(sys.argv) > 1:
        p = Path(sys.argv[1])
        if not p.exists():
            sys.exit(f"❌ Arquivo não encontrado: {p}")
        return p

    pasta = Path(CONFIG["pasta_rh"])
    padroes = ["rela*-*.xlsx", "rela*.*.xlsx", "efetivo*.xlsx"]
    candidatos = []
    for p in padroes:
        candidatos.extend(pasta.glob(p))
    # Dedup e ordena por mtime desc
    candidatos = sorted(set(candidatos),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidatos:
        sys.exit(f"❌ Nenhum arquivo de relação/efetivo encontrado em {pasta}")
    return candidatos[0]


def extrair_data(arquivo: Path) -> str:
    """Aceita DD-MM, DD.MM ou DD_MM (com ou sem ano). Normaliza para DD-MM."""
    m = re.search(r"(\d{2})[-._](\d{2})(?:[-._](\d{2,4}))?", arquivo.stem)
    if not m:
        return datetime.now().strftime("%d-%m")
    dd, mm, yy = m.group(1), m.group(2), m.group(3)
    return f"{dd}-{mm}-{yy}" if yy else f"{dd}-{mm}"


def main():
    pasta_hist = Path(CONFIG["pasta_historico"])
    pasta_hist.mkdir(parents=True, exist_ok=True)

    arquivo_novo = resolver_arquivo_novo()
    data_str = extrair_data(arquivo_novo)
    print(f"📂 Processando: {arquivo_novo.name}  (data: {data_str})")

    df_novo = carregar_planilha(arquivo_novo)
    n_pj  = sum(1 for k in df_novo.index if k.startswith("PJ::"))
    n_clt = sum(1 for k in df_novo.index if k.startswith("CLT::"))
    print(f"   {len(df_novo)} colaboradores ({n_clt} CLT, {n_pj} PJ)")

    snapshots = sorted(pasta_hist.glob("snapshot_*.xlsx"))
    snapshot_anterior = snapshots[-1] if snapshots else None
    consolidado = pasta_hist / CONFIG["arquivo_consolidado"]

    if snapshot_anterior is None:
        print("ℹ️  Nenhum snapshot anterior — esta será a baseline.")
        wb = load_workbook(consolidado) if consolidado.exists() else Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]
        nome_aba = proximo_nome_aba(wb, data_str)
        ws = wb.create_sheet(nome_aba)
        ws.merge_cells("A1:E1")
        c = ws["A1"]
        c.value = f"Baseline inicial - {data_str}"
        c.font, c.fill, c.alignment = FONT_TITLE, FILL_TITLE, CENTER
        ws["A3"] = (f"Primeira execução. {len(df_novo)} colaboradores registrados "
                    "como referência. A próxima execução já mostrará as mudanças.")
        ws["A3"].font = FONT_BODY
        ws.merge_cells("A3:E3")
        ws.column_dimensions["A"].width = 30
    else:
        print(f"📊 Comparando com: {snapshot_anterior.name}")
        df_antigo = carregar_planilha(snapshot_anterior)
        adicoes, remocoes, alteracoes = comparar(df_antigo, df_novo)
        print(f"   ➕ {len(adicoes)} adições")
        print(f"   ➖ {len(remocoes)} remoções")
        print(f"   ✏️  {len(alteracoes)} alterações")

        # Diagnóstico: se "tudo mudou", provavelmente é problema de chave
        total_antigo = len(df_antigo)
        if total_antigo > 0 and len(remocoes) / total_antigo > 0.5:
            print(f"\n⚠️  ALERTA: mais de 50% dos colaboradores foram marcados como removidos.")
            print("   Isso geralmente indica problema de chave (formato de matrícula diferente).")
            print("   Amostra de chaves no snapshot antigo:")
            for k in list(df_antigo.index)[:5]:
                print(f"     {k}")
            print("   Amostra de chaves na planilha nova:")
            for k in list(df_novo.index)[:5]:
                print(f"     {k}")
            print()

        wb = load_workbook(consolidado) if consolidado.exists() else Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]
        escrever_aba(wb, data_str, arquivo_novo.name, adicoes, remocoes, alteracoes)

    atualizar_indice(wb)
    wb.save(consolidado)
    print(f"💾 Consolidado: {consolidado}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    snapshot_path = pasta_hist / f"snapshot_{timestamp}_{data_str}.xlsx"
    shutil.copy2(arquivo_novo, snapshot_path)
    print(f"📸 Snapshot salvo: {snapshot_path.name}\n✅ Pronto.")


if __name__ == "__main__":
    main()